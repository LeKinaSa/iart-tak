from tak import State
import openpyxl

def test_negamax(state, depth, cuts, caching, n_try):
    state.negamax(depth, cuts, caching, True)

    file = "Statistics.xlsx"
    workbook = openpyxl.load_workbook(file)
    sheet = workbook["Negamax"]
    n_rows = sheet.max_row
    line_value = ""
    if cuts:
        line_value += "T"
    else:
        line_value += "F"
    if caching:
        line_value += "T"
    else:
        line_value += "F"
    line_value += str(depth)

    line = None
    column = n_try + 2
    for index in range (1, n_rows + 1):
        value = sheet.cell(index, 1).value
        if value == line_value:
            line = index
            break
    if line == None:
        raise Exception(line_value + " was not found")

    sheet.cell(line    , column).value = State.nm_calls
    sheet.cell(line + 1, column).value = State.nm_prunings
    sheet.cell(line + 2, column).value = State.nm_cache_hits
    sheet.cell(line + 3, column).value = State.nm_time_possible_moves
    sheet.cell(line + 4, column).value = State.nm_time_evaluating
    
    workbook.save(file)


for n in range(1, 11):
    for depth in range(1, 6):
        test_negamax(State(4), depth,  True,  True, n)
        if depth <= 4:
            test_negamax(State(4), depth, False,  True, n)
            test_negamax(State(4), depth,  True, False, n)
            test_negamax(State(4), depth, False, False, n)

